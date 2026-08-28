"""Export des simulations vers CSV / Excel et génération du résumé texte."""

from __future__ import annotations

import csv
from pathlib import Path
from typing import TYPE_CHECKING

from utils.currency import (
    decimal_to_str,
    format_dzd,
    format_money,
    format_percent,
)

if TYPE_CHECKING:  # évite une dépendance circulaire à l'exécution
    from models.simulation import Simulation

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    HAS_OPENPYXL = True
except ImportError:  # pragma: no cover - dépend de l'installation
    HAS_OPENPYXL = False


class ExportError(Exception):
    """Erreur lors d'un export CSV / Excel."""


HEADERS = [
    "Date",
    "Marque",
    "Modèle",
    "Année",
    "Devise",
    "Cylindrée (L)",
    "Prix véhicule",
    "Fret",
    "Taux de change",
    "Prix véhicule (DZD)",
    "Fret (DZD)",
    "Valeur douanière (DZD)",
    "Taux douane (%)",
    "Droits de douane (DZD)",
    "Base TVA (DZD)",
    "TVA (DZD)",
    "Frais transitaire (DZD)",
    "Frais portuaires (DZD)",
    "Taxe véhicule (DZD)",
    "Coût total (DZD)",
]

_NUMERIC_COLUMNS = list(range(5, 20))  # colonnes F..T numériques (devise = texte)


def _row_from_sim(sim: Simulation) -> list:
    return [
        sim.date.strftime("%d/%m/%Y"),
        sim.marque,
        sim.modele,
        sim.annee,
        sim.devise,
        float(sim.cylindree),
        float(sim.prix_usd),
        float(sim.fret_usd),
        float(sim.taux_change),
        float(sim.prix_dzd),
        float(sim.fret_dzd),
        float(sim.valeur_douaniere),
        float(sim.taux_douane),
        float(sim.droits_douane),
        float(sim.base_tva),
        float(sim.tva),
        float(sim.frais_transitaire),
        float(sim.frais_portuaires),
        float(sim.taxe_vehicule),
        float(sim.cout_total),
    ]


def _csv_safe_text(value: str) -> str:
    """Neutralise les débuts de cellule risqués à l'ouverture dans Excel
    (« = », « + », « - », « @ », tabulation) : injection de formules CSV."""
    return "'" + value if value[:1] in ("=", "+", "-", "@", "\t") else value


def _csv_cell(value):
    """Prépare une cellule CSV : texte sécurisé, entier brut, nombre français."""
    if isinstance(value, str):
        return _csv_safe_text(value)
    if isinstance(value, int):
        return value
    return _csv_number(value)


def _csv_number(value: float) -> str:
    """Nombre au format français pour Excel (séparateur « ; » + virgule)."""
    text = f"{value:.2f}".rstrip("0").rstrip(".")
    return text.replace(".", ",")


def export_simulations_csv(path: str | Path, sims: list[Simulation]) -> Path:
    """Exporte les simulations au format CSV (séparateur « ; », UTF-8 BOM)."""
    path = Path(path)
    try:
        with open(path, "w", newline="", encoding="utf-8-sig") as handle:
            writer = csv.writer(handle, delimiter=";")
            writer.writerow(HEADERS)
            for sim in sims:
                writer.writerow([_csv_cell(value) for value in _row_from_sim(sim)])
    except OSError as exc:
        raise ExportError(f"Impossible d'écrire le fichier CSV : {exc}") from exc
    return path


def export_simulations_excel(path: str | Path, sims: list[Simulation]) -> Path:
    """Exporte les simulations au format Excel (.xlsx)."""
    if not HAS_OPENPYXL:
        raise ExportError(
            "Le module openpyxl est requis pour l'export Excel.\n"
            "Installez-le avec : pip install openpyxl"
        )
    path = Path(path)
    try:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Simulations"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="14345C")
        for col, title in enumerate(HEADERS, start=1):
            cell = sheet.cell(row=1, column=col, value=title)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for r, sim in enumerate(sims, start=2):
            for c, value in enumerate(_row_from_sim(sim), start=1):
                cell = sheet.cell(row=r, column=c, value=value)
                if c in _NUMERIC_COLUMNS:
                    cell.number_format = "#,##0.00"

        widths = [12, 14, 16, 8, 8, 11, 15, 13, 12, 17, 14, 18, 12, 17, 16, 14, 16, 16, 14, 17]
        for c, width in enumerate(widths, start=1):
            sheet.column_dimensions[get_column_letter(c)].width = width
        sheet.freeze_panes = "A2"
        workbook.save(path)
    except OSError as exc:
        raise ExportError(f"Impossible d'écrire le fichier Excel : {exc}") from exc
    return path


def simulation_summary_text(sim: Simulation) -> str:
    """Résumé textuel prêt à copier dans le presse-papiers."""
    lines = [
        f"{sim.marque.upper()} {sim.modele.upper()} {sim.annee}",
        f"Cylindrée : {decimal_to_str(sim.cylindree)} L",
        f"Prix véhicule : {format_money(sim.prix_usd, sim.devise)}",
        f"Fret : {format_money(sim.fret_usd, sim.devise)}",
        f"Taux : {decimal_to_str(sim.taux_change)} DA/USD",
        "",
        f"Prix véhicule : {format_dzd(sim.prix_dzd)}",
        f"Fret : {format_dzd(sim.fret_dzd)}",
        f"Valeur douanière : {format_dzd(sim.valeur_douaniere)}",
        f"Droits de douane ({format_percent(sim.taux_douane / 100)}) : {format_dzd(sim.droits_douane)}",
        f"Base TVA : {format_dzd(sim.base_tva)}",
        f"TVA : {format_dzd(sim.tva)}",
        f"Transitaire : {format_dzd(sim.frais_transitaire)}",
        f"Frais portuaires : {format_dzd(sim.frais_portuaires)}",
        f"Taxe véhicule : {format_dzd(sim.taxe_vehicule)}",
        "",
        f"Coût total estimé : {format_dzd(sim.cout_total)}",
    ]
    return "\n".join(lines)
